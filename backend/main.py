from fastapi import FastAPI, HTTPException, Query, Request, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.templating import Jinja2Templates
from datetime import datetime
from uuid import uuid4
from pymongo import MongoClient
from openpyxl import Workbook
from weasyprint import HTML
from jinja2 import Environment, FileSystemLoader
import os

from schema import RequirementRequest, RequirementResponse
from utils import build_search_filter, calculate_storage, estimate_bitrate, recommend_server, update_bitrate


app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

client = MongoClient("mongodb://192.168.1.67:27017")
db = client["redx_vms"]
collection = db["requirements"]
EXPORT_DIR = "exports"
TEMPLATE_DIR = "templates"
os.makedirs(EXPORT_DIR, exist_ok=True)
env = Environment(loader=FileSystemLoader(TEMPLATE_DIR))

templates = Jinja2Templates(directory=TEMPLATE_DIR)



@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    return templates.TemplateResponse("index1.html", {"request": request})


@app.post("/requirement", response_model=RequirementResponse)
def create_requirement(req: RequirementRequest):
    try:
        update_bitrate(req)

        uid = str(uuid4())

        for cam in req.camera_configs:
            if cam.bitrate_kbps:
                bitrate_mbps = cam.bitrate_kbps / 1000
            else:
                bitrate_mbps = estimate_bitrate(cam.resolution, cam.fps, cam.codec)         
            total_bitrate = bitrate_mbps * cam.qty

        print("TOtal Bitrate:", total_bitrate)

        max_retention = max(c.retention_days for c in req.camera_configs)
        avg_record_hour = max(c.record_hour for c in req.camera_configs)
        camera_configs = [cam.dict() for cam in req.camera_configs]
        doc = {
            "_id": uid,
            "customer_name": req.customer_name,
            "project_name": req.project_name,
            "location": req.location,
            "assigned_person": req.assigned_person,
            "camera_configs": camera_configs,
            "created_at": datetime.utcnow(),
            "bandwidth": round(total_bitrate, 2),
            "storage_tb": calculate_storage(total_bitrate, max_retention, avg_record_hour),   #This is for Recording Space not OS 
            "server_spec": recommend_server(sum(cam.qty for cam in req.camera_configs), total_bitrate, round(total_bitrate, 2), max_retention, avg_record_hour, camera_configs)
        }
        collection.insert_one(doc)
        return RequirementResponse(**doc, id=uid)

    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/requirement/{id}", response_model=RequirementResponse)
def get_requirement(id: str):
    doc = collection.find_one({"_id": id})
    if not doc:
        raise HTTPException(status_code=404, detail="Requirement not found")
    return RequirementResponse(**doc, id=doc.pop("_id"))


@app.get("/requirement/{id}/export/xlsx")
def export_excel(id: str):
    doc = collection.find_one({"_id": id})
    if not doc:
        raise HTTPException(status_code=404, detail="Requirement not found")

    wb = Workbook()
    ws = wb.active
    ws.title = "REDX VMS Requirement"

    ws.append(["Customer", doc["customer_name"]])
    ws.append(["Project", doc["project_name"]])
    ws.append(["Location", doc["location"]])
    ws.append(["Assigned Person", doc["assigned_person"]])
    ws.append([])
    ws.append(["Camera Brand", "Resolution", "FPS", "Codec", "Record Hour", "Retention Days", "Quantity"])
    for cam in doc["camera_configs"]:
        ws.append([
            cam["name"], cam["resolution"], cam["fps"], cam["codec"],
            cam["record_hour"], cam["retention_days"], cam["qty"]
        ])
    ws.append([])
    ws.append(["Estimated Bitrate (Mbps)", doc["bandwidth"]])
    ws.append(["Estimated Storage (TB)", doc["storage_tb"]])
    ws.append(["Server CPU", doc["server_spec"]["cpu"]])
    ws.append(["RAM (GB)", doc["server_spec"]["ram_gb"]])
    ws.append(["HDD Total (TB)", doc["server_spec"]["hdd_tb"]])
    ws.append(["NIC Ports", doc["server_spec"]["nic"]])

    path = os.path.join(EXPORT_DIR, f"requirement_{id}.xlsx")
    wb.save(path)
    return FileResponse(path, filename=f"redx_report_{id}.xlsx")


@app.get("/requirement/{id}/export/pdf")
def export_pdf(id: str):
    doc = collection.find_one({"_id": id})

    if not doc:
        raise HTTPException(status_code=404, detail="Requirement not found")
    
    for cam in doc.get("camera_configs"):
        quantity = cam.get('qty')
        bitrate_kbps = cam.get('bitrate_kbps')
        cam["bandwidth"] = round(bitrate_kbps * quantity / 1024, 2)
    print(type(doc), doc)

    template = env.get_template("report_template.html")
    html_out = template.render(data=doc, created=datetime.now().strftime("%d-%b-%Y"))

    pdf_path = os.path.join(EXPORT_DIR, f"requirement_{id}.pdf")
    HTML(string=html_out).write_pdf(pdf_path)
    return FileResponse(pdf_path, filename=f"redx_report_{id}.pdf")


@app.get("/requirement/list/")
def list_all_requirements():
    results = []
    for doc in collection.find().sort("created_at", -1):
        total_qty = sum(cam.get("qty", 1) for cam in doc.get('camera_configs', []))
        total_bitrate = doc.get('bandwidth', 0)
        total_bandwidth = round(total_bitrate * total_qty / 1024, 2)

        results.append({
            "id": str(doc["_id"]),
            "project_name": doc["project_name"],
            "customer_name": doc["customer_name"],
            "location": doc.get("location", ""),
            "assigned_person": doc.get("assigned_person", ""),
            "created_at": doc.get("created_at", "")
        })
    return results


@app.get("/requirement/export/all/xlsx")
def export_all_excel():
    docs = list(collection.find())
    if not docs:
        raise HTTPException(status_code=404, detail="No requirements found")

    wb = Workbook()
    ws = wb.active
    ws.title = "All REDX Requirements"

    ws.append([
        "Customer", "Project", "Location", "Assigned", "Bitrate (Mbps)",
        "Storage (TB)", "Camera Brand", "Resolution", "FPS", "Codec",
        "Record Hour", "Retention Days", "Qty"
    ])

    for doc in docs:
        for cam in doc["camera_configs"]:
            ws.append([
                doc["customer_name"],
                doc["project_name"],
                doc["location"],
                doc["assigned_person"],
                doc["bandwidth"],
                doc["storage_tb"],
                cam["name"],
                cam["resolution"],
                cam["fps"],
                cam["codec"],
                cam["record_hour"],
                cam["retention_days"],
                cam["qty"]
            ])

    path = os.path.join(EXPORT_DIR, f"all_requirements.xlsx")
    wb.save(path)
    return FileResponse(path, filename=f"redx_all_requirements.xlsx")



@app.get("/requirement/search/")
def search_requirements(
    query: str = Query(None, description="Search term to match across fields"),
    customer_name: str = Query(None, description="Filter by customer name"),
    project_name : str = Query(None, description='Filter by project name'),
    location: str = Query(None, description="Filter by location"),
    assigned_person : str = Query(None, description="Filter by assigned person"),
    start_date : datetime = Query(None, description="Start date created_at filter"),
    end_date : datetime = Query(None, description="End date at created at filter"),
):
    search_filter = build_search_filter(
        query=query,
        customer_name=customer_name,
        project_name=project_name,
        location=location,
        assigned_person=assigned_person,
        start_date=start_date,
        end_date=end_date
    )
    print("search_filter:", search_filter)
    results = []
    for doc in collection.find(search_filter).sort("created_at", -1):
        results.append({
            "id": str(doc["_id"]),
            "project_name": doc["project_name"],
            "customer_name": doc["customer_name"],
            "location": doc.get("location", ""),
            "assigned_person": doc.get("assigned_person", ""),
            "created_at": doc.get("created_at", "")
        })

    return {
        "count": len(results),
        "results": results
    }



@app.get("/requirement/export/filtered/xlsx")
def export_filtered_excel(
    query: str = Query(None, description="Search term to match across fields"),
    customer_name: str = Query(None, description="Filter by customer name"),
    project_name: str = Query(None, description="Filter by project name"),
    location: str = Query(None, description="Filter by location"),
    assigned_person: str = Query(None, description="Filter by assigned person"),
    start_date: datetime = Query(None, description="Start date for created_at filter"),
    end_date: datetime = Query(None, description="End date for created_at filter")
):
    search_filter = build_search_filter(
        query=query,
        customer_name=customer_name,
        project_name=project_name,
        location=location,
        assigned_person=assigned_person,
        start_date=start_date,
        end_date=end_date
    )

    docs = list(collection.find(search_filter).sort("created_at", -1))
    if not docs:
        raise HTTPException(status_code=404, detail="No requirements found matching the filters")

    wb = Workbook()
    ws = wb.active
    ws.title = "Filtered REDX Requirements"

    ws.append(["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    if query:
        ws.append(["Search Query", query])
    if customer_name:
        ws.append(["Customer Filter", customer_name])
    if project_name:
        ws.append(["Project Filter", project_name])
    if location:
        ws.append(["Location Filter", location])
    if assigned_person:
        ws.append(["Assigned Person Filter", assigned_person])
    if start_date or end_date:
        date_range = []
        if start_date:
            date_range.append(f"From: {start_date}")
        if end_date:
            date_range.append(f"To: {end_date}")
        ws.append(["Date Range", " ".join(date_range)])
    ws.append([])

    ws.append([
        "Customer", "Project", "Location", "Assigned", "Created At",
        "Camera Brand", "Resolution", "FPS", "Codec", "Record Hour", 
        "Retention Days", "Qty", "Bitrate (Mbps)", "Storage (TB)"
    ])

    for doc in docs:
        for cam in doc.get("camera_configs", []):
            created_at = doc.get("created_at", "")
            if isinstance(created_at, datetime):
                created_at = created_at.strftime("%Y-%m-%d %H:%M:%S")
            
            ws.append([
                doc["customer_name"],
                doc["project_name"],
                doc.get("location", ""),
                doc.get("assigned_person", ""),
                created_at,
                cam["name"],
                cam["resolution"],
                cam["fps"],
                cam["codec"],
                cam["record_hour"],
                cam["retention_days"],
                cam["qty"],
                doc.get("bandwidth", "N/A"),
                doc.get("storage_tb", "N/A")
            ])

    filename_parts = ["redx_export"]
    if customer_name:
        filename_parts.append(f"cust_{customer_name[:20]}")
    if project_name:
        filename_parts.append(f"proj_{project_name[:20]}")
    if query:
        filename_parts.append(f"q_{query[:10]}")
    filename = "_".join(filename_parts) + ".xlsx"
    path = os.path.join(EXPORT_DIR, filename)    
    os.makedirs(EXPORT_DIR, exist_ok=True)
    wb.save(path)
    
    return FileResponse(path, filename=filename)


# @app.get("/requirement/export/all/xlsx")
# def export_all_excel():
#     docs = list(collection.find())
#     if not docs:
#         raise HTTPException(status_code=404, detail="No requirements found")

#     wb = Workbook()
#     ws = wb.active
#     ws.title = "All REDX Requirements"

#     ws.append([
#         "Customer", "Project", "Location", "Assigned", "Bitrate (Mbps)",
#         "Storage (TB)", "Camera Brand", "Resolution", "FPS", "Codec",
#         "Record Hour", "Retention Days", "Qty"
#     ])

#     for doc in docs:
#         for cam in doc["camera_configs"]:
#             ws.append([
#                 doc["customer_name"],
#                 doc["project_name"],
#                 doc["location"],
#                 doc["assigned_person"],
#                 doc["bandwidth"],
#                 doc["storage_tb"],
#                 cam["name"],
#                 cam["resolution"],
#                 cam["fps"],
#                 cam["codec"],
#                 cam["record_hour"],
#                 cam["retention_days"],
#                 cam["qty"]
#             ])

#     path = os.path.join(EXPORT_DIR, f"all_requirements.xlsx")
#     wb.save(path)
#     return FileResponse(path, filename=f"redx_all_requirements.xlsx")


if __name__ == "__main__":
    
    import uvicorn
    uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=True)
